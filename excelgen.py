def make_excel(list, sheetname, office):
	import openpyxl,datetime
	from openpyxl.styles import Alignment, Border, Side

	now = datetime.datetime.now()
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = sheetname

	

	sh = wb[sheetname]
	sh.column_dimensions['A'].width = 60
	sh.column_dimensions['B'].width = 40
	sh.column_dimensions['C'].width = 30
	sh.column_dimensions['I'].width = 15
	sh.column_dimensions['F'].width = 15
	sh.column_dimensions['K'].width = 15
	columns = ["物件url", "物件名", "住所", "家賃", "管理費", "月々支払い", "敷金", "礼金", "仲介手数料", "面積", "１m2の値段", "初期費用", "距離"]
	sh.append(columns)

	for cells in sh['1']:
		cells.alignment = Alignment(horizontal = 'center')
		cells.border = Border(
			outline=True,
			bottom=Side(style="thick", color="FF000000")
			)

	for i in list:
		sh.append(i)

	
	for cells in sh['A']:
		currow = cells.row
		if cells.row != 1:
			initialcost = sh.cell(row=currow, column=6).value *2 + sh.cell(row=currow, column=5).value +sh.cell(row=currow, column=6).value + sh.cell(row=currow, column=7).value 
			sh.cell(row=currow, column=12).value = initialcost

	
 #get the distance from website
	import urllib.parse 
	from bs4 import BeautifulSoup
	from selenium import webdriver
	import time
	
	driver = webdriver.PhantomJS()

	for loc in sh['C']:
		if loc.value != "住所":
			locinfo = urllib.parse.quote_plus(loc.value)
			base =f"http://r1web.realwork.jp/index_ex.html?{urllib.parse.quote_plus(f'{office}',encoding='utf-8')}&{locinfo}"
			driver.get(base)
			
			time.sleep(2)
			data = driver.page_source.encode('utf-8')
			


			soup = BeautifulSoup(data, "lxml")
			div = soup.find_all("td", {'id':'meter_r'})
			div = str(div)
			div = div.replace('[<td id="meter_r" style="width: 50mm;border-style: solid;border-color: #E0E0E0">', '')
			div = div.replace('</td>]', '')
			
			print(div)
			
			currow = loc.row
			sh.cell(row=currow, column=13).value= div
	wb.save(f"media/{sheetname}_{now}.xlsx")









